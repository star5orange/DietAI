#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel测试用例生成器 - 使用openpyxl库

使用方法:
    python generate_test_cases.py <output_path> [case_count]

示例:
    python generate_test_cases.py ./output.xlsx
    python generate_test_cases.py ./output.xlsx 50
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 列宽配置
COLUMN_WIDTHS = {
    'A': 8,    # 序号
    'B': 28,   # 测试点标题
    'C': 16,   # 所属模块
    'D': 18,   # 二级模块
    'E': 30,   # 前置条件
    'F': 45,   # 操作步骤
    'G': 35,   # 测试数据
    'H': 45,   # 预期结果
    'I': 15,   # 实际结果
    'J': 20,   # 备注
    'K': 10,   # 版本
    'L': 10,   # 是否废止
}

DEFAULT_ROW_HEIGHT = 50
HEADER_ROW_HEIGHT = 25


def create_styles():
    """创建样式"""
    return {
        'header_font': Font(bold=True, size=11),
        'header_fill': PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'cell_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'cell_alignment': Alignment(vertical='top', wrap_text=True),
    }


def generate_test_cases_template():
    """生成空的测试用例模板"""
    return [
        ["序号", "测试点标题", "所属模块", "二级模块", "前置条件", "操作步骤", "测试数据", "预期结果", "实际结果", "备注", "版本", "是否废止"],
    ]


def create_workbook(test_cases=None, output_path=None):
    """
    创建Excel工作簿

    Args:
        test_cases: 测试用例列表，每一项是一个包含12个字段的列表
        output_path: 输出文件路径

    Returns:
        Workbook对象
    """
    if test_cases is None:
        test_cases = generate_test_cases_template()

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    styles = create_styles()
    headers = test_cases[0]

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['cell_border']
        cell.alignment = styles['header_alignment']

    # 写入数据
    for row_idx, row_data in enumerate(test_cases[1:], 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['cell_border']
            cell.alignment = styles['cell_alignment']

    # 设置列宽
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # 设置行高
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    for row_idx in range(2, len(test_cases) + 1):
        ws.row_dimensions[row_idx].height = DEFAULT_ROW_HEIGHT

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存文件
    if output_path:
        wb.save(output_path)
        print(f"[OK] Test cases file created: {output_path}")
        print(f"     Total cases: {len(test_cases) - 1}")

    return wb


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("Usage: python generate_test_cases.py <output_path> [case_count]")
        sys.exit(1)

    output_path = sys.argv[1]
    case_count = int(sys.argv[2]) if len(sys.argv) > 2 else 0

    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 生成空模板或带示例数据的文件
    test_cases = generate_test_cases_template()

    # 如果指定了用例数量，生成示例数据
    if case_count > 0:
        for i in range(1, case_count + 1):
            test_cases.append([
                str(i),
                f"测试点标题_{i}",
                "所属模块",
                "二级模块",
                "前置条件",
                "操作步骤",
                "测试数据",
                "预期结果",
                "",
                "",
                "V1.0",
                "否"
            ])

    create_workbook(test_cases, output_path)


if __name__ == "__main__":
    main()
